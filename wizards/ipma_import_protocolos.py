import base64
from datetime import datetime, date
from io import BytesIO
import unicodedata

from odoo import models, fields
from odoo.exceptions import UserError


class IpmaImportProtocolosWizard(models.TransientModel):
	_name = 'ipma.import.protocolos.wizard'
	_description = 'Importar Protocolos (XLSX)'

	file = fields.Binary(string='Ficheiro XLSX', required=True)
	filename = fields.Char(string='Nome do ficheiro')

	def action_import(self):
		self.ensure_one()
		if not self.file:
			raise UserError('Selecione um ficheiro XLSX para importar.')

		try:
			from openpyxl import load_workbook
		except Exception as exc:
			raise UserError('A biblioteca openpyxl nao esta instalada no servidor.') from exc

		data = base64.b64decode(self.file)
		workbook = load_workbook(filename=BytesIO(data), data_only=True)
		sheet = workbook.active

		header_row = None
		for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
			header_row = list(row)
			break

		if not header_row:
			raise UserError('O ficheiro nao contem cabecalho.')

		mapping = self._build_header_mapping(header_row)
		if not mapping:
			raise UserError('Nenhuma coluna reconhecida no cabecalho.')

		protocol_model = self.env['ipma.protocolo']
		for row in sheet.iter_rows(min_row=2, values_only=True):
			vals = {}
			for idx, field_name in mapping.items():
				value = row[idx] if idx < len(row) else None
				parsed = self._parse_cell_value(field_name, value)
				if parsed is not None:
					vals[field_name] = parsed

			if vals:
				protocol_model.create(vals)

		return {'type': 'ir.actions.act_window_close'}

	def _build_header_mapping(self, header_row):
		field_labels = [
			('siged', 'SIGED'),
			('tipo_documento', 'Tipo de documento'),
			('ambito_geografico', 'Ambito geografico'),
			('entidades_signatarias', 'Entidades signatarias'),
			('classificacao_entidades', 'Classificacao da(s) entidade(s)'),
			('designacao_documento', 'Designacao do documento'),
			('sigla', 'Sigla'),
			('ambito_objeto', 'Ambito / Objeto'),
			('data_assinatura', 'Data de assinatura'),
			('entrada_vigor', 'Entrada em vigor'),
			('data_vigencia', 'Data de Vigencia'),
			('periodo_vigencia', 'Periodo de Vigencia'),
			('antecedencia_revisao_denuncia', 'Antecedencia para revisao e denuncia'),
			('obrigacoes_financeiras', 'Obrigacoes financeiras'),
			('valor', 'Valor'),
			('ponto_focal_ipma_nome', 'Ponto Focal IPMA - Nome'),
			('ponto_focal_ipma_uo', 'Ponto Focal IPMA - UO'),
			('ponto_focal_ipma_tlf', 'Ponto Focal IPMA - Tlf.'),
			('ponto_focal_ipma_email', 'Ponto Focal IPMA - email'),
			('ponto_focal_parceiros_nome', 'Ponto Focal (parceiros) - Nome'),
			('ponto_focal_parceiros_tlf', 'Ponto Focal (parceiros) - Tlf.'),
			('ponto_focal_parceiros_email', 'Ponto Focal (parceiros) - email'),
			('outros_protocolos_mesmas_entidades', 'Outros protocolos com a(s) mesma(s) entidade(s)'),
			('numero_documento_relacionado', 'N.o Documento Relacionado'),
			('ativo', 'Ativo'),
			('observacoes', 'Observacoes'),
		]

		normalized = {}
		for field_name, label in field_labels:
			normalized[self._normalize_header(field_name)] = field_name
			normalized[self._normalize_header(label)] = field_name

		mapping = {}
		for idx, raw in enumerate(header_row):
			if raw is None:
				continue
			key = self._normalize_header(str(raw))
			field_name = normalized.get(key)
			if field_name:
				mapping[idx] = field_name

		return mapping

	def _normalize_header(self, value):
		cleaned = value.strip().lower()
		cleaned = cleaned.replace('\n', ' ').replace('\t', ' ')
		cleaned = ' '.join(cleaned.split())
		cleaned = cleaned.replace('/', ' ').replace('-', ' ').replace('.', ' ')
		cleaned = cleaned.replace('(', ' ').replace(')', ' ')
		cleaned = cleaned.replace(':', ' ').replace(';', ' ')
		cleaned = cleaned.replace(',', ' ')
		cleaned = cleaned.replace('  ', ' ')
		cleaned = cleaned.replace(' ', '_')
		cleaned = unicodedata.normalize('NFKD', cleaned).encode('ascii', 'ignore').decode('ascii')
		return cleaned

	def _parse_cell_value(self, field_name, value):
		if value is None or value == '':
			return None

		if field_name in {'data_assinatura', 'entrada_vigor', 'data_vigencia'}:
			if isinstance(value, datetime):
				return value.date()
			if isinstance(value, date):
				return value
			if isinstance(value, str):
				parsed = self._parse_date_string(value)
				if parsed:
					return parsed
			return None

		if field_name == 'valor':
			if isinstance(value, str):
				try:
					return float(value.replace(',', '.'))
				except Exception:
					return None
			if isinstance(value, (int, float)):
				return float(value)
			return None

		if field_name == 'ativo':
			if isinstance(value, bool):
				return value
			if isinstance(value, (int, float)):
				return bool(int(value))
			if isinstance(value, str):
				value_norm = value.strip().lower()
				return value_norm in {'1', 'true', 'sim', 'yes', 'y', 'x'}
			return None

		return str(value).strip()

	def _parse_date_string(self, value):
		value = value.strip()
		for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
			try:
				return datetime.strptime(value, fmt).date()
			except Exception:
				continue
		return None


